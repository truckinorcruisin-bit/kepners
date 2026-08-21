"""Synthetic validation of market_drift.

Fixture rules learned the hard way:
  1. WAR must be INDEPENDENT of 2026 ADP, or talent drift cancels price drift by
     construction and the test proves nothing.
  2. The slot index must be assigned IN RANK ORDER within each position, or an
     injected tier-wide shift scatters across unrelated slots and the smoother
     averages it away.
  3. Scenarios must run in ISOLATION. Talent drift is a *relative* measure --
     collapsing one position genuinely raises every other position's WAR rank --
     so a confound case run alongside the control cases contaminates them.

Scenario A (price only):  QB1-8 cheaper by 12, RB25-45 pricier by 10, no talent
                          change      -> TruRank should track price, WR neutral
Scenario B (confound):    TE1-10 cheaper by 12 AND that tier's talent collapses
                          -> TruRank must NOT read as a bargain
"""
import json, os, random
import openpyxl
import market_drift as md

POS_COUNT = {"QB": 32, "RB": 65, "WR": 75, "TE": 28}
REPL = {"QB": 15, "RB": 30, "WR": 36, "TE": 15}
YAHOO_LEAN = 4.0   # Yahoo sits a constant ~4 picks off the blended average
STATS = "espn_season_stats_2025.json"


def build(seed, shifts, talent_hit=None, yahoo25=False, yahoo26=False):
    """shifts/talent_hit: fn(pos, slot_index) -> picks / points delta."""
    random.seed(seed)
    pool = [p for p, n in POS_COUNT.items() for _ in range(n)]
    random.shuffle(pool)
    seen, board_rows = {}, []
    for rank, pos in enumerate(pool, start=1):
        i = seen.get(pos, 0); seen[pos] = i + 1
        board_rows.append((pos, i, float(rank)))

    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "2025 Big Board"
    HDR = 6
    cells = [(7, "Player"), (8, "Team"), (9, "Pos"), (21, "Avg Rank")]
    if yahoo25:
        cells.append((13, "Yahoo"))
    for c, lbl in cells:
        ws.cell(HDR, c, lbl)
    for r, (pos, i, rk) in enumerate(board_rows, start=HDR + 1):
        ws.cell(r, 7, f"P25 {pos}{i+1}"); ws.cell(r, 8, "KC")
        ws.cell(r, 9, pos); ws.cell(r, 21, rk)
        if yahoo25:
            # Yahoo leans a constant offset off the blend, plus its own jitter
            ws.cell(r, 13, rk + YAHOO_LEAN + random.uniform(-2, 2))
    board = md.read_2025_board(wb)

    quality = {(pos, i): 320 - rk * 0.85 for pos, i, rk in board_rows}

    stats = {"players_by_position": {}}
    for pos, i, rk in board_rows:
        stats["players_by_position"].setdefault(pos, []).append(
            {"name": f"P25 {pos}{i+1}",
             "total_points": max(0.0, quality[(pos, i)] + random.gauss(0, 12))})
    with open(STATS, "w") as f:
        json.dump(stats, f)

    keyed = sorted((rk + shifts(pos, i) + random.uniform(-1.2, 1.2), pos, i)
                   for pos, i, rk in board_rows)

    pts26 = {}
    for pos, i, rk in board_rows:
        q = quality[(pos, i)] - (talent_hit(pos, i) if talent_hit else 0)
        pts26[(pos, i)] = max(0.0, q + random.gauss(0, 6))

    by_pos = {}
    for (pos, i), pts in pts26.items():
        by_pos.setdefault(pos, []).append(pts)
    repl_pts = {}
    for pos, plist in by_pos.items():
        plist.sort(reverse=True)
        repl_pts[pos] = plist[min(REPL[pos], len(plist)) - 1]

    players = []
    for newrank, (_, pos, i) in enumerate(keyed, start=1):
        war = round(pts26[(pos, i)] - repl_pts[pos], 1)
        rec = {"id": f"p-{pos}{i+1}", "name": f"P26 {pos}{i+1}",
               "pos": pos, "avgRank": float(newrank),
               "projectedWar": war, "warByLeague": {"kepners": war}}
        if yahoo26:
            rec["platform"] = {"yahoo": float(newrank) + YAHOO_LEAN + random.uniform(-2, 2)}
        players.append(rec)

    info = md.compute(players, board, ["kepners"], {"kepners": REPL},
                      platform_by_league={"kepners": "yahoo"} if yahoo26 else None)
    return ({p["name"]: p["marketDrift"]["kepners"] for p in players}, board,
            players, info.get("kepners", {}))


def mean(xs): return sum(xs) / len(xs)
def show(lbl, grp):
    print(f"  {lbl:10s} price={mean([d['price'] for d in grp]):6.1f}  "
          f"talent={mean([d['talent'] for d in grp]):6.1f}  "
          f"tru={mean([d['tru'] for d in grp]):6.1f}  "
          f"bands={sorted({d['band'] for d in grp})}")


# ---------------- Scenario A: price moves only ------------------------------
print("\n== Scenario A: QB tier cheaper, RB dead zone pricier, talent flat ==")
D, board, players, _ = build(11, lambda p, i: (+12 if (p == "QB" and i < 8) else
                                            -10 if (p == "RB" and 24 <= i < 45) else 0))
qb = [D[f"P26 QB{i}"] for i in range(2, 8)]
rb = [D[f"P26 RB{i}"] for i in range(28, 42)]
wr = [D[f"P26 WR{i}"] for i in range(10, 40)]
show("QB2-7", qb); show("RB28-41", rb); show("WR10-39", wr)
assert 6 <= mean([d["tru"] for d in qb]) <= 18, mean([d["tru"] for d in qb])
assert -18 <= mean([d["tru"] for d in rb]) <= -6, mean([d["tru"] for d in rb])
assert abs(mean([d["tru"] for d in wr])) < 6
# Individual slots inside a shifted tier can still land in the neutral deadband;
# what must never happen is a slot coloured the WRONG WAY.
assert not any(d["band"] in ("neg", "strong-neg") for d in qb), "QB miscoloured"
assert not any(d["band"] in ("pos", "strong-pos") for d in rb), "RB miscoloured"
assert sum(d["band"] in ("pos", "strong-pos") for d in qb) >= len(qb) * 0.6
assert sum(d["band"] in ("neg", "strong-neg") for d in rb) >= len(rb) * 0.6
assert all(d["band"] == "neutral" for d in wr), "control WRs must stay uncoloured"
print("PASS: both injected price trends recovered; control position uncoloured")

# ---------------- Scenario B: the confound ----------------------------------
print("\n== Scenario B: TE tier cheaper by 12 BUT talent collapses ==")
D2, _, _, _ = build(11, lambda p, i: (+12 if (p == "TE" and i < 10) else 0),
                 talent_hit=lambda p, i: (60 if (p == "TE" and i < 10) else 0))
te = [D2[f"P26 TE{i}"] for i in range(2, 9)]
show("TE2-8", te)
assert mean([d["price"] for d in te]) > 7, "price move should still be visible"
assert mean([d["talent"] for d in te]) < -8, "talent collapse should be detected"
assert mean([d["tru"] for d in te]) < 4, "TruRank must not read as a bargain"
assert not any(d["band"] == "strong-pos" for d in te)
print("PASS: CONFOUND CAUGHT — cheap on price alone, TruRank refuses the bait")

# ---------------- degradation paths -----------------------------------------
print("\n== Degradation ==")
os.remove(STATS)
p2 = [dict(p) for p in players]
md.compute(p2, board, ["kepners"], {"kepners": REPL})
assert {x["marketDrift"]["kepners"]["sample"] for x in p2} <= {
    "price-only", "beyond-2025-depth", "n/a"}
coloured = [x for x in p2 if x["marketDrift"]["kepners"]["band"] not in ("neutral", "none")]
assert coloured, "price-only mode must still derive bands, not go all-neutral"
print(f"PASS: no 2025 stats -> price-only, bands still derived ({len(coloured)} coloured)")

p3 = [dict(p) for p in players]
assert md.compute(p3, [], ["kepners"], {"kepners": REPL}) == {}
assert all(x["marketDrift"] == {} for x in p3)
print("PASS: no 2025 board -> feature disabled cleanly")

wb2 = openpyxl.Workbook(); wb2.active.title = "Nope"
assert md.read_2025_board(wb2) == []
print("PASS: missing 2025 tab degrades quietly")

assert md._band(3.0, {"strongPos": 2, "pos": 1, "neg": -1, "strongNeg": -2}) == "neutral"
assert md._band(None, None) == "none"
print("PASS: deadband + null handling")

# ---------------- league-specific ADP source --------------------------------
print("\n== ADP source routing ==")
shifts = lambda p, i: (+12 if (p == "QB" and i < 8) else 0)

Dy, _, _, infoY = build(11, shifts, yahoo25=True, yahoo26=True)
assert infoY["adpSource2026"] == "yahoo" and infoY["adpSource2025"] == "yahoo"
assert infoY["likeForLike"] is True
qy = [Dy[f"P26 QB{i}"] for i in range(2, 8)]
assert 6 <= mean([d["tru"] for d in qy]) <= 18, mean([d["tru"] for d in qy])
print(f"  yahoo->yahoo  like-for-like={infoY['likeForLike']}  "
      f"QB2-7 tru={mean([d['tru'] for d in qy]):.1f}")
print("PASS: platform-to-platform comparison cancels the constant Yahoo lean")

Dm, _, _, infoM = build(11, shifts, yahoo25=False, yahoo26=True)
assert infoM["adpSource2026"] == "yahoo" and infoM["adpSource2025"] == "average"
assert infoM["likeForLike"] is False
qm = [Dm[f"P26 QB{i}"] for i in range(2, 8)]
wm = [Dm[f"P26 WR{i}"] for i in range(10, 40)]
print(f"  yahoo->average  like-for-like={infoM['likeForLike']}  "
      f"QB2-7 tru={mean([d['tru'] for d in qm]):.1f}  "
      f"WR10-39 tru={mean([d['tru'] for d in wm]):.1f}")
# The mixed-source run carries the constant lean, but the SHAPE must survive:
# the shifted QB tier still has to stand well clear of the untouched WRs.
assert mean([d["tru"] for d in qm]) - mean([d["tru"] for d in wm]) > 6, "shape lost"
assert abs(mean([d["tru"] for d in wm])) < 6, "constant lean not removed from level"
assert infoM.get("leanRemoved") is not None, "lean correction should have fired"
print("PASS: mixed sources flagged, lean removed, positional shape recovered")

Ds, _, _, infoS = build(11, shifts, yahoo25=True, yahoo26=False)
assert infoS["adpSource2026"] == "average", infoS
print("PASS: no 2026 platform column -> falls back to blended average")

# scratch rows / #N/A ranks
wb3 = openpyxl.Workbook(); ws3 = wb3.active; ws3.title = "2025 Big Board"
for c, lbl in [(2, "Player"), (3, "Team"), (4, "Pos"), (9, "Avg Rank")]:
    ws3.cell(4, c, lbl)
for r in range(5, 15):
    ws3.cell(r, 2, f"Real Guy {r}"); ws3.cell(r, 3, "KC")
    ws3.cell(r, 4, "WR"); ws3.cell(r, 9, float(r))
for r in range(20, 30):
    ws3.cell(r, 2, f"scratch {r}")
ws3.cell(35, 2, "Bad Rank"); ws3.cell(35, 3, "KC"); ws3.cell(35, 4, "WR")
ws3.cell(35, 9, "#N/A")
got = md.read_2025_board(wb3)
assert len(got) == 10, len(got)
print("PASS: header autodetect on a shifted row; scratch + #N/A excluded")
print("\nALL DRIFT TESTS PASSED")
