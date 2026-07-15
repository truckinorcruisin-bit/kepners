#!/usr/bin/env python3
"""
convert_bigboard.py
Converts the FY Big Board Excel into the JSON the Mission Control draft app consumes.

Usage:
    python convert_bigboard.py FY26_BigBoard.xlsx data/bigboard.json

Output schema:
{
  "meta": { "season": 2026, "generated": "..." },
  "players": [
     { "id", "name", "team", "pos", "seanPosRank", "tier", "tierGroup",
       "targetRound", "like", "avgRank", "avgRound", "myDiff",
       "platform": {"yahoo":..,"espn":..,"underdog":..,...}, "notes" }
  ],
  "leagues": {
     "kepners": { "platform":"yahoo", "myTeam":"Pickups", "teams":[...], "rosterSlots":[...],
                  "keepers":[...], "rules": {see league_rules.json} },
     "miami":   { "platform":"yahoo", "myTeam":"HannahLees", ... },
     "zimmer":  { "platform":"espn",  "myTeam":"Elements of Intrigue", ... }
  }
}

League format/keeper rules live in league_rules.json (hand-maintained, not
derived from the spreadsheet) and are merged into each league's output here.
Edit that file directly to update rules; this script only merges it in.
"""
import sys, os, json, re, glob
from datetime import datetime, timezone
import openpyxl

RULES_FILE = "league_rules.json"  # hand-maintained; sits alongside this script

# Same replacement-level assumption used in zimmer_draft_grades.py for grading
# past drafts -- kept in sync here (not imported, since this script has no
# other dependency on that module) so projected WAR uses identical logic to
# historical WAR. If you tune one, tune the other.
REPLACEMENT_RANK = {
    "QB": 15, "RB": 30, "WR": 36, "TE": 15, "K": 12, "DEF": 12, "D/ST": 12,
}

SUFFIX_RE = re.compile(r"\b(jr|sr|ii|iii|iv|v)\.?\b", re.IGNORECASE)
PUNCT_RE = re.compile(r"[.\'\-]")


def normalize_name(name):
    """Lowercase, strip Jr./Sr./numeral suffixes and punctuation, collapse
    whitespace -- used to join Big Board players (from Sean's Excel) against
    ESPN's player pool (from espn_player_values.py), since the Big Board has
    no ESPN player_id to join on directly. Best-effort: nicknames or heavily
    reformatted names may still miss; unmatched players are logged."""
    if not name:
        return ""
    n = name.lower()
    n = PUNCT_RE.sub("", n)
    n = SUFFIX_RE.sub("", n)
    n = re.sub(r"\s+", " ", n).strip()
    return n


def load_espn_player_values():
    """Loads the most recent espn_player_values_<year>.json if present.
    Returns (values_by_normalized_name, replacement_by_position) or ({}, {})
    if no such file exists yet."""
    candidates = sorted(glob.glob("espn_player_values_*.json"), reverse=True)
    if not candidates:
        print("Note: no espn_player_values_*.json found -- players will have "
              "no espnRecommendedBid/projectedWar until that's pulled.")
        return {}, {}

    with open(candidates[0]) as f:
        data = json.load(f)
    print(f"Loaded {candidates[0]} ({len(data['players'])} players).")

    values_by_name = {}
    by_pos = {}
    for p in data["players"]:
        values_by_name[normalize_name(p["name"])] = p
        by_pos.setdefault(p["position"], []).append(p)

    replacement_by_position = {}
    for pos, plist in by_pos.items():
        plist.sort(key=lambda x: x.get("projected_total_points") or 0, reverse=True)
        rank = REPLACEMENT_RANK.get(pos, 20)
        idx = min(rank, len(plist)) - 1
        if idx >= 0:
            replacement_by_position[pos] = plist[idx].get("projected_total_points") or 0
    return values_by_name, replacement_by_position


def merge_espn_values(players):
    """Attaches espnRecommendedBid / projectedPoints / projectedWar to each Big
    Board player by normalized-name match. Fields are left null if unmatched or
    if no espn_player_values file exists -- the site's Player Card shows
    '—' gracefully in that case, this never fails the build."""
    values_by_name, replacement_by_position = load_espn_player_values()
    if not values_by_name:
        for p in players:
            p["espnRecommendedBid"] = None
            p["projectedPoints"] = None
            p["projectedWar"] = None
            p["byeWeek"] = None
            p["proTeamEspn"] = None
        return

    unmatched = []
    for p in players:
        key = normalize_name(p["name"])
        ev = values_by_name.get(key)
        if not ev:
            unmatched.append(p["name"])
            p["espnRecommendedBid"] = None
            p["projectedPoints"] = None
            p["projectedWar"] = None
            p["byeWeek"] = None
            p["proTeamEspn"] = None
            continue
        proj = ev.get("projected_total_points")
        replacement = replacement_by_position.get(p["pos"])
        p["espnRecommendedBid"] = ev.get("auction_value_avg")
        p["projectedPoints"] = proj
        p["byeWeek"] = ev.get("bye_week")
        p["proTeamEspn"] = ev.get("pro_team")
        p["projectedWar"] = (
            round(proj - replacement, 1) if (proj is not None and replacement is not None) else None
        )

    if unmatched:
        print(f"WARNING: {len(unmatched)} Big Board players had no ESPN name match "
              f"(showing first 15): {unmatched[:15]}")


def load_league_rules():
    """Static format/keeper-rule config, hand-maintained in league_rules.json.
    Merged into each league's output under the "rules" key. Missing file is
    non-fatal -- the app just won't have rule metadata until it's added."""
    if not os.path.exists(RULES_FILE):
        print(f"Note: {RULES_FILE} not found -- leagues will have no 'rules' section.")
        return {}
    with open(RULES_FILE) as f:
        rules = json.load(f)
    rules.pop("_comment", None)
    return rules


def norm(s):
    return re.sub(r"\s+", " ", str(s)).strip() if s is not None else None


def slug(name):
    return re.sub(r"[^a-z0-9]+", "-", str(name).lower()).strip("-")


def tier_group(tier):
    """Collapse the verbose tier string into a coarse actionable band."""
    if not tier or tier == "#N/A":
        return "unranked"
    t = tier.lower()
    if "don't draft" in t or "don\u2019t draft" in t or "streamer" in t or "replacement" in t:
        return "avoid"
    if "handcuff" in t:
        return "handcuff"
    if "late round" in t or "very late" in t:
        return "late-flier"
    m = re.match(r"([A-Z]+)\d", tier)
    return m.group(1).lower() if m else "ranked"


def read_players(wb):
    ws = wb["2025 Big Board"]  # rename tab per season; keep logic identical
    # Column map (1-indexed) confirmed from header row 6
    COL = dict(targetRound=4, like=6, player=7, team=8, pos=9, tier=10,
               seanPosRank=11, yahoo=13, underdog=14, cbs=16, espn=17,
               ffpc=18, sleeper=19, nfl=20, avgRank=21, avgRound=22,
               myDiff=23, notes=25)
    players = []
    for r in range(7, ws.max_row + 1):
        name = norm(ws.cell(r, COL["player"]).value)
        if not name:
            continue
        pos = norm(ws.cell(r, COL["pos"]).value)
        if pos in (None, "#N/A"):
            pos = "NA"
        tier = norm(ws.cell(r, COL["tier"]).value)
        players.append({
            "id": slug(name),
            "name": name,
            "team": norm(ws.cell(r, COL["team"]).value),
            "pos": pos,
            "seanPosRank": ws.cell(r, COL["seanPosRank"]).value,
            "tier": tier,
            "tierGroup": tier_group(tier),
            "targetRound": ws.cell(r, COL["targetRound"]).value,
            "like": norm(ws.cell(r, COL["like"]).value),
            "avgRank": ws.cell(r, COL["avgRank"]).value,
            "avgRound": ws.cell(r, COL["avgRound"]).value,
            "myDiff": ws.cell(r, COL["myDiff"]).value,
            "platform": {
                "yahoo": ws.cell(r, COL["yahoo"]).value,
                "espn": ws.cell(r, COL["espn"]).value,
                "underdog": ws.cell(r, COL["underdog"]).value,
                "cbs": ws.cell(r, COL["cbs"]).value,
                "ffpc": ws.cell(r, COL["ffpc"]).value,
                "sleeper": ws.cell(r, COL["sleeper"]).value,
                "nfl": ws.cell(r, COL["nfl"]).value,
            },
            "notes": norm(ws.cell(r, COL["notes"]).value),
        })
    return players


def read_team_sheet(wb, sheet, platform, my_team):
    """Parse a *Team* sheet: managers, draft slots, roster template, keepers."""
    ws = wb[sheet]
    # Team names live on row 3 starting col E; managers on row 4
    teams = []
    col = 5
    while True:
        tname = norm(ws.cell(3, col).value)
        if not tname:
            break
        teams.append({
            "slot": col - 4,
            "team": tname,
            "manager": norm(ws.cell(4, col).value),
            "isMe": my_team.lower() in (tname or "").lower(),
        })
        col += 1

    # Roster template: col D rows 5+ ('QB','RB',...)
    roster_slots = []
    r = 5
    while True:
        slot = norm(ws.cell(r, 4).value)
        if not slot:
            break
        roster_slots.append(slot)
        r += 1

    # Keepers: cells in the team grid formatted "Player (round)"
    keepers = []
    for tr in range(5, 5 + len(roster_slots)):
        for tc in range(5, 5 + len(teams)):
            cell = norm(ws.cell(tr, tc).value)
            if not cell:
                continue
            m = re.match(r"(.+?)\s*\((\d+)\)\s*$", cell)
            if m:
                keepers.append({
                    "team": teams[tc - 5]["team"],
                    "player": norm(m.group(1)),
                    "playerId": slug(m.group(1)),
                    "round": int(m.group(2)),
                })
    return {
        "platform": platform,
        "myTeam": my_team,
        "teams": teams,
        "rosterSlots": roster_slots,
        "keepers": keepers,
    }


def main(src, dst):
    wb = openpyxl.load_workbook(src, data_only=True)
    rules = load_league_rules()
    players = read_players(wb)
    merge_espn_values(players)
    out = {
        "meta": {"season": 2026, "generated": datetime.now(timezone.utc).isoformat()},
        "players": players,
        "leagues": {
            "kepners": read_team_sheet(wb, "Kepners Team", "yahoo", "Pickups"),
            "miami": read_team_sheet(wb, "Miami Team", "yahoo", "HannahLees"),
            # Zimmer (ESPN) has no dedicated Team sheet in FY25 — stub for now,
            # will populate once the FY26 workbook adds it.
            "zimmer": {"platform": "espn", "myTeam": "Elements of Intrigue",
                       "teams": [], "rosterSlots": [], "keepers": []},
        },
    }
    for league_key, league_data in out["leagues"].items():
        if league_key in rules:
            league_data["rules"] = rules[league_key]

    with open(dst, "w") as f:
        json.dump(out, f, indent=2)
    print(f"Wrote {dst}: {len(out['players'])} players; "
          f"leagues={list(out['leagues'])}")
    for k, v in out["leagues"].items():
        has_rules = "rules" in v
        print(f"  {k}: {len(v['teams'])} teams, {len(v['keepers'])} keepers, "
              f"roster={v['rosterSlots']}, rules_loaded={has_rules}")


if __name__ == "__main__":
    src = sys.argv[1] if len(sys.argv) > 1 else "/mnt/user-data/uploads/FY25_BigBoard.xlsx"
    dst = sys.argv[2] if len(sys.argv) > 2 else "bigboard.json"
    main(src, dst)
